import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Title of the app
st.title("Excel Sheet Viewer with Formulas")

# File upload
uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file:
    # Load the uploaded Excel file using openpyxl
    workbook = load_workbook(uploaded_file, data_only=False)  # Keep formulas

    # Display all sheet names
    st.write("**Sheet Names:**")
    sheet_names = workbook.sheetnames
    st.write(sheet_names)

    # Loop through each sheet, clean it, and display values with formulas
    for sheet_name in sheet_names:
        st.write(f"### Sheet: {sheet_name}")
        
        # Select the sheet
        sheet = workbook[sheet_name]
        
        # Collect data with formulas
        data = []
        for row in sheet.iter_rows(values_only=False):  # Access cells, not just values
            row_data = []
            for cell in row:
                # Check if the cell contains a formula
                if cell.data_type == 'f':  # 'f' indicates a formula
                    row_data.append(f"= {cell.value}")  # Display formula
                else:
                    row_data.append(cell.value if cell.value is not None else "")  # Replace None with empty string
            data.append(row_data)

        # Create a DataFrame for display
        df = pd.DataFrame(data, columns=[str(col[0].value) for col in sheet.iter_cols(1, sheet.max_column, 1, 1)])
        
        # Remove 'Unnamed: 0' column if it exists
        if 'Unnamed: 0' in df.columns:
            df = df.drop(columns=['Unnamed: 0'])
        
        # Convert the DataFrame to an HTML table and display it
        st.markdown(df.to_html(index=False), unsafe_allow_html=True)
